import streamlit as st
import os
import re
import time
import hashlib
import io
import fitz           # PyMuPDF
import docx
import pandas as pd
import gspread
import openpyxl
from google.oauth2.service_account import Credentials
import google.generativeai as genai
from google.api_core.exceptions import ResourceExhausted
import openai
from supabase import create_client, Client

# ==============================================================================
# CONFIGURACIÓN DE PÁGINA
# ==============================================================================
st.set_page_config(
    page_title="Bot de Bitácora | DESPUX",
    page_icon="📋",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ==============================================================================
# CONFIGURACIÓN SUPABASE (siempre visible, no es secreto de IA)
# ==============================================================================
SUPABASE_URL = "https://rromxmhmadwtshughttz.supabase.co"
SUPABASE_KEY = "sb_publishable_GAusBaYKd_ED7Vl1_k7VRA_MZI8BEwu"

# ==============================================================================
# HELPERS DE SECRETS (Streamlit Cloud → st.secrets, local → os.getenv)
# ==============================================================================
def get_secret(key, default=None):
    try:
        return st.secrets[key]
    except:
        return os.getenv(key, default)

def get_all_api_keys():
    return (
        get_secret("GEMINI_API_KEY"),
        get_secret("GEMINI_API_KEY_BACKUP"),
        get_secret("DEEPSEEK_API_KEY"),
        get_secret("SPREADSHEET_NAME"),
    )

# ==============================================================================
# SESSION STATE — equivalente a los atributos de clase en la versión tkinter
# ==============================================================================
defaults = {
    "logged_in_user": None,
    "api_1_enabled": True,   # Gemini Principal
    "api_2_enabled": True,   # Gemini Backup
    "api_3_enabled": True,   # DeepSeek
    "datos_procesados": [],
    "current_api_index": 0,
    "processing_done": False,
    "log_lines": [],
    "show_login_modal": False,
    "show_cred_modal": False,
    "admin_unlocked": False,
}
for k, v in defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ==============================================================================
# ESTILOS CSS PERSONALIZADOS (ESTILO APPLE / V28 REFOZADO)
# ==============================================================================
st.markdown("""
<style>
    /* Forzar fondo blanco y texto oscuro global */
    .stApp, [data-theme="light"], [data-theme="dark"] { 
        background-color: #F8F9FA !important; 
        color: #1C1C1E !important; 
    }
    
    /* Sidebar Robusto */
    div[data-testid="stSidebar"] { 
        background-color: #FFFFFF !important; 
        border-right: 1px solid #E5E5EA !important;
        min-width: 300px !important;
    }
    
    /* Asegurar visibilidad de textos en sidebar */
    div[data-testid="stSidebar"] p, div[data-testid="stSidebar"] div, div[data-testid="stSidebar"] span {
        color: #1C1C1E !important;
    }
    
    /* Títulos Sidebar */
    .sb-heading { 
        font-family: 'Helvetica', sans-serif; 
        font-size: 22px; 
        font-weight: bold; 
        color: #1C1C1E !important; 
        margin-bottom: 20px; 
    }
    
    /* Logo DESPUX */
    .despux-logo { 
        font-family: 'Helvetica', sans-serif; 
        font-size: 32px; 
        font-weight: 900; 
        color: #002366 !important; 
        text-decoration: none !important; 
        display: block; 
        text-align: center; 
        padding: 20px 0; 
    }

    /* Cards (Contenedores Blancos) - Usando selectores de atributo más estables */
    div[data-testid="stVerticalBlock"] > div[style*="border: 1px solid"] {
        background-color: white !important;
        border-radius: 12px !important;
        padding: 20px !important;
        box-shadow: 0 4px 6px rgba(0,0,0,0.05) !important;
    }
    
    /* Botones - Forzar visibilidad del texto */
    button p { 
        color: white !important; 
        font-weight: bold !important; 
    }
    
    /* Botones específicos por color */
    button[key="btn_procesar"], button[key="btn_login"] { background-color: #007AFF !important; }
    button[key="btn_detener"], button[key="btn_limpiar"] { background-color: #FF3B30 !important; }
    button[key="btn_generar"], div[data-testid="stDownloadButton"] > button { background-color: #FF9500 !important; }
    
    /* Terminal de Sistema */
    .terminal-card { 
        background-color: #F2F2F7 !important; 
        border-radius: 12px !important; 
        padding: 15px !important; 
        border: 1px solid #E5E5EA !important;
        font-family: 'Consolas', monospace !important;
        height: 250px;
        overflow-y: auto;
        color: #1C1C1E !important;
    }

    /* Inputs y Textareas */
    [data-testid="stWidgetLabel"] p { color: #1C1C1E !important; font-weight: 600 !important; }
    input, textarea { 
        background-color: #FFFFFF !important; 
        color: #1C1C1E !important; 
        border: 1px solid #D1D1D6 !important; 
    }
    
    /* Ocultar elementos molestos */
    #MainMenu, footer, header { visibility: hidden !important; }
</style>
""", unsafe_allow_html=True)

# ==============================================================================
# FUNCIONES AUXILIARES
# ==============================================================================
def hash_password(password: str) -> str:
    return hashlib.sha256(password.encode()).hexdigest()

def log(msg: str, level: str = "normal"):
    st.session_state.log_lines.append((msg, level))

def get_gsheets_client():
    """Crea cliente de Google Sheets usando secrets de Streamlit o archivo JSON local."""
    scopes = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
    try:
        # Streamlit Cloud: credenciales embebidas en secrets
        creds_dict = dict(st.secrets["gcp_service_account"])
        creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
    except:
        # Local: archivo JSON
        json_file = get_secret("GOOGLE_SHEETS_CREDENTIALS_FILE", "bitacora-491012-71d4c0e7a86f.json")
        creds = Credentials.from_service_account_file(json_file, scopes=scopes)
    return gspread.authorize(creds)

def inicializar_modelo_gemini(api_key: str):
    genai.configure(api_key=api_key)
    try:
        disponibles = [m.name for m in genai.list_models() if 'generateContent' in m.supported_generation_methods]
        prioridades = ['models/gemini-2.5-flash', 'models/gemini-2.0-flash', 'models/gemini-1.5-flash-latest', 'models/gemini-1.5-flash']
        nombre = next((p.replace('models/', '') for p in prioridades if p in disponibles), 'gemini-1.5-flash')
    except:
        nombre = 'gemini-1.5-flash'
    return genai.GenerativeModel(nombre), nombre

def consultar_ia(modelo, prompt: str, reintentos_maximos=4):
    """Consulta la IA con fallback Gemini → Backup → DeepSeek (respetando los flags habilitados)."""
    GEMINI_KEY, GEMINI_BACKUP_KEY, DEEPSEEK_KEY, _ = get_all_api_keys()

    for intento in range(reintentos_maximos):
        api_index = st.session_state.current_api_index
        try:
            if api_index < 2:
                respuesta = modelo.generate_content(prompt)
                return respuesta.text.strip(), modelo
            else:
                client = openai.OpenAI(api_key=DEEPSEEK_KEY, base_url="https://api.deepseek.com")
                resp = client.chat.completions.create(
                    model="deepseek-chat",
                    messages=[
                        {"role": "system", "content": "Eres un ingeniero perito experto rellenando planillas."},
                        {"role": "user", "content": prompt}
                    ],
                    temperature=0.2
                )
                return resp.choices[0].message.content.strip(), modelo
        except Exception as e:
            err_str = str(e).lower()
            api_2_ok = GEMINI_BACKUP_KEY and st.session_state.api_2_enabled
            api_3_ok = DEEPSEEK_KEY and st.session_state.api_3_enabled

            if "quota" in err_str or "429" in err_str or isinstance(e, ResourceExhausted):
                if api_index == 0 and api_2_ok:
                    st.session_state.current_api_index = 1
                    log("🚨 Límite API Gemini Principal. Cambiando a Backup...", "error")
                    genai.configure(api_key=GEMINI_BACKUP_KEY)
                    modelo, _ = inicializar_modelo_gemini(GEMINI_BACKUP_KEY)
                elif api_index == 0 and not api_2_ok and api_3_ok:
                    st.session_state.current_api_index = 2
                    log("🔥 Backup desactivado. Saltando a DeepSeek...", "error")
                elif api_index <= 1 and api_3_ok:
                    st.session_state.current_api_index = 2
                    log("🔥 Límite Gemini. Cambiando a DeepSeek...", "error")
                else:
                    espera = 20 * (intento + 1)
                    log(f"⚠️ Todas las APIs agotadas. Esperando {espera}s...", "error")
                    time.sleep(espera)
            else:
                log(f"❌ Error IA: {e}", "error")
                time.sleep(5)
    return None, modelo

def parsear_respuesta(respuesta: str, coords_requeridas: list) -> dict:
    """Parser robusto con regex + fallback línea por línea para respuestas de Gemini/DeepSeek."""
    resultados = {}
    patron = re.compile(
        r'(?:^\*{0,2})(B\d+)\*{0,2}\s*:\s*(.+?)(?=\n\*{0,2}B\d+|\Z)',
        re.MULTILINE | re.DOTALL
    )
    for match in patron.finditer(respuesta):
        coord = match.group(1).strip().upper()
        texto = re.sub(r'\*+', '', match.group(2).strip())
        if coord in coords_requeridas and coord not in resultados:
            resultados[coord] = texto

    # Fallback si regex no capturó nada
    if not resultados:
        for linea in respuesta.split("\n"):
            if ":" in linea:
                try:
                    l = linea.replace("**", "").replace("*", "").strip()
                    partes = l.split(":", 1)
                    coord = partes[0].strip().upper()
                    texto = partes[1].strip()
                    if coord in coords_requeridas and coord not in resultados:
                        resultados[coord] = texto
                except:
                    continue
    return resultados

def extraer_texto_pdf(pdf_bytes: bytes) -> str:
    texto = ""
    with fitz.open(stream=pdf_bytes, filetype="pdf") as doc:
        for page in doc:
            texto += page.get_text()
    return texto

def extraer_texto_extras(archivos) -> str:
    texto = ""
    for archivo in archivos:
        try:
            if archivo.name.endswith(".txt"):
                texto += f"\n--- {archivo.name} ---\n" + archivo.read().decode("utf-8", errors="ignore")
            elif archivo.name.endswith(".docx"):
                d = docx.Document(io.BytesIO(archivo.read()))
                texto += f"\n--- {archivo.name} ---\n" + "\n".join(p.text for p in d.paragraphs)
        except Exception as e:
            log(f"⚠️ No se pudo leer {archivo.name}: {e}", "error")
    return texto

# ==============================================================================
# FUNCIÓN PRINCIPAL DE PROCESAMIENTO
# ==============================================================================
def procesar_todo(pdf_bytes, archivos_extras, datos_usuario, progress_bar, lbl_progreso, tabla_placeholder, log_placeholder):
    """Ejecuta el flujo completo: extracción → Sheets → IA → resultados."""
    st.session_state.current_api_index = 0
    st.session_state.datos_procesados = []
    st.session_state.log_lines = []

    GEMINI_KEY, _, _, SPREADSHEET_NAME = get_all_api_keys()

    log(">> Extrayendo texto del PDF...")
    pdf_texto = extraer_texto_pdf(pdf_bytes)
    log(f"   ({len(pdf_texto)} caracteres extraídos del PDF)")

    texto_extra = ""
    if archivos_extras:
        texto_extra = extraer_texto_extras(archivos_extras)
        log(f">> {len(archivos_extras)} archivos extras procesados.")

    log(">> Conectando a Google Sheets...")
    try:
        cli = get_gsheets_client()
    except Exception as e:
        log(f"[ERROR] No se pudo conectar a Google Sheets: {e}", "error")
        return

    log(">> Configurando API de Gemini...")
    try:
        modelo, nombre = inicializar_modelo_gemini(GEMINI_KEY)
        log(f"   Usando modelo: {nombre}")
    except Exception as e:
        log(f"[ERROR] No se pudo inicializar Gemini: {e}", "error")
        return

    # Pre-limpieza de Columna B
    log(">> [Pre-proceso] Limpiando Columna B en Google Sheets...")
    try:
        doc_pre = cli.open(SPREADSHEET_NAME)
        for h in doc_pre.worksheets():
            h.batch_clear(["B2:B5000"])
        log("   ✅ Columna B limpiada.", "success")
    except Exception as e:
        log(f"   [Aviso] No se pudo limpiar: {e}", "error")

    # Iterar hojas
    try:
        documento = cli.open(SPREADSHEET_NAME)
        pestanas = documento.worksheets()
    except Exception as e:
        log(f"[ERROR] No existe el documento '{SPREADSHEET_NAME}': {e}", "error")
        return

    contextos = [f"- {k}: {v}" for k, v in datos_usuario.items() if v]
    bloque_usuario = "\n".join(contextos) if contextos else "Ninguno adicional."

    for hoja in pestanas:
        log(f"\n--- Analizando pestaña: {hoja.title} ---")
        try:
            matriz = hoja.get_all_values()
        except Exception as e:
            log(f"  [ERROR] No se pudo leer la hoja: {e}", "error")
            continue

        if not matriz:
            continue

        datos_a_procesar = []
        for r_idx, fila in enumerate(matriz):
            col_a = fila[0].strip() if len(fila) > 0 else ""
            col_b = fila[1].strip() if len(fila) > 1 else ""
            if col_a and not col_b:
                coord = gspread.utils.rowcol_to_a1(r_idx + 1, 2)
                datos_a_procesar.append({"etiqueta": col_a, "coordenada": coord, "hoja": hoja.title})

        if not datos_a_procesar:
            log("  No hay celdas vacías en Columna B.")
            continue

        total_filas = len(datos_a_procesar)
        log(f"  Encontradas {total_filas} filas a rellenar.")
        lbl_progreso.text(f"Procesando: 0/{total_filas} celdas")
        progress_bar.progress(0)
        filas_procesadas = 0

        texto_oferta_seguro = pdf_texto[:50000]
        texto_extra_seguro = texto_extra[:20000]
        texto_info = f"--- 2. OFERTA TÉCNICA PRINCIPAL (PDF) ---\n{texto_oferta_seguro}\n"
        if texto_extra_seguro:
            texto_info += f"\n--- 3. NOTAS/ARCHIVOS EXTRAS ---\n{texto_extra_seguro}\n"

        lote_tamano = 3
        for i in range(0, total_filas, lote_tamano):
            lote = datos_a_procesar[i:i + lote_tamano]
            coords_requeridas = [item["coordenada"] for item in lote]

            intentos_lote = 0
            max_intentos_lote = 3
            lote_ok = False

            while intentos_lote < max_intentos_lote and not lote_ok:
                intentos_lote += 1
                lote_str = "\n".join(f"{item['coordenada']}: {item['etiqueta']}" for item in lote)

                prompt = (
                    f"Actúa como un procesador de datos analítico y estricto. Tu tarea es evaluar la información de la 'Columna A' y generar la respuesta correspondiente en la 'Columna B'.\n\n"
                    f"Fuentes de información provistas:\n"
                    f"--- 1. DATOS MANUALES DEL USUARIO ---\n{bloque_usuario}\n\n"
                    f"{texto_info}\n"
                    f"-----------------------------------------\n\n"
                    f"REGLAS OBLIGATORIAS PARA CADA FILA:\n"
                    f"1. PROCESAMIENTO TOTAL (CERO VACÍOS): Debes generar una respuesta para TODAS las celdas de la lista. Ninguna puede quedar en blanco.\n"
                    f"2. REGLA DE CONTINGENCIA ('sin informacion'): Si el dato no existe o la etiqueta no tiene sentido, responde exactamente: sin informacion.\n"
                    f"3. PROHIBIDO DUPLICAR: Nunca copies el texto de la Columna A en la Columna B.\n"
                    f"4. EXHAUSTIVIDAD TOTAL (PROHIBIDO RESUMIR): Si la etiqueta pide 'descripción detallada', 'hitos', 'etapas', 'fases', o cualquier tipo de desglose:\n"
                    f"   - Enumera y describe TODOS y CADA UNO de los hitos, etapas o fases. No solo el primero.\n"
                    f"   - Si el hito tuvo 4 etapas, describe las 4. Si tuvo 6, describe las 6. NUNCA truncar.\n"
                    f"   - Usa viñetas o números para separar cada etapa claramente.\n\n"
                    f"CAMPOS A COMPLETAR (Formato COORDENADA: ETIQUETA):\n"
                    f"{lote_str}\n\n"
                    f"FORMATO ESTRICTO DE SALIDA:\n"
                    f"COORDENADA: DATO_INFERIDO\n"
                )

                if intentos_lote > 1:
                    log(f"  [Re-intento {intentos_lote}] Lote incompleto anterior...")

                log(f"  >> Consultando IA (Lote {i//lote_tamano + 1}, Intento {intentos_lote})...")
                respuesta, modelo = consultar_ia(modelo, prompt)

                if respuesta:
                    resultados_lote = parsear_respuesta(respuesta, coords_requeridas)
                    updates_batch = []

                    for coord, texto in resultados_lote.items():
                        etiqueta = next(x['etiqueta'] for x in lote if x['coordenada'] == coord)
                        log(f"    -> {coord} = {texto[:60]}...")
                        st.session_state.datos_procesados.append({
                            "Celda": coord,
                            "Col A (Etiqueta)": etiqueta,
                            "Col B (Generado)": texto,
                            "Hoja": hoja.title
                        })
                        updates_batch.append({'range': coord, 'values': [[texto]]})

                    if len(resultados_lote) >= len(lote):
                        lote_ok = True
                        filas_procesadas += len(resultados_lote)
                        progreso = filas_procesadas / max(total_filas, 1)
                        progress_bar.progress(min(progreso, 1.0))
                        lbl_progreso.text(f"Procesando: {filas_procesadas}/{total_filas} celdas")

                        if updates_batch:
                            try:
                                hoja.batch_update(updates_batch)
                                log(f"  ✅ [LOTE COMPLETADO] {len(updates_batch)} celdas actualizadas.", "success")
                            except Exception as e_up:
                                log(f"  [Aviso] batch_update falló, usando celda por celda: {e_up}", "error")
                                for d in updates_batch:
                                    try:
                                        hoja.update_acell(d['range'], d['values'][0][0])
                                        time.sleep(0.5)
                                    except:
                                        pass
                    else:
                        log(f"  [!] Lote incompleto ({len(resultados_lote)}/{len(lote)}). Reintentando...")
                else:
                    log("  [ERROR] La IA no respondió.", "error")
                    break

                # Actualizar tabla en vivo
                if st.session_state.datos_procesados:
                    tabla_placeholder.dataframe(
                        pd.DataFrame(st.session_state.datos_procesados),
                        use_container_width=True,
                        height=250
                    )
                # Actualizar log en vivo
                render_log(log_placeholder)

    progress_bar.progress(1.0)
    lbl_progreso.text("✅ ¡Proceso completado!")
    log("\n[ÉXITO] 🎉 Tarea terminada. Todas las celdas han sido procesadas.", "success")
    st.session_state.processing_done = True
    render_log(log_placeholder)

def render_log(placeholder):
    """Renderiza el log con colores en un área de texto."""
    html = "<div style='background:#F2F2F7; padding:10px; border-radius:8px; height:220px; overflow-y:auto; font-family:Consolas,monospace; font-size:12px;'>"
    for msg, level in st.session_state.log_lines:
        color = "#34C759" if level == "success" else ("#FF3B30" if level == "error" else "#1C1C1E")
        safe = msg.replace("<", "&lt;").replace(">", "&gt;").replace("\n", "<br>")
        html += f"<span style='color:{color}'>{safe}</span><br>"
    html += "</div>"
    placeholder.markdown(html, unsafe_allow_html=True)

# ==============================================================================
# MÓDULO DE AUTENTICACIÓN (Sidebar)
# ==============================================================================
def sidebar_auth():
    if st.session_state.logged_in_user:
        u = st.session_state.logged_in_user
        st.sidebar.success(f"👤 {u['nombre']} {u['apellido']}")
        st.sidebar.caption(u['empresa'])
        if st.sidebar.button("🚪 Cerrar Sesión", use_container_width=True):
            st.session_state.logged_in_user = None
            st.session_state.processing_done = False
            st.rerun()
    else:
        if st.sidebar.button("👤 Iniciar Sesión", use_container_width=True, key="btn_login"):
            st.session_state.show_login_modal = True
            st.rerun()

@st.dialog("Acceso de Usuario", width="small")
def modal_login():
    tab_login, tab_registro = st.tabs(["Iniciar Sesión", "Registrarse"])

    with tab_login:
        email = st.text_input("Correo Electrónico", placeholder="usuario@empresa.com", key="login_email")
        passwd = st.text_input("Contraseña", type="password", key="login_pass")
        if st.button("Ingresar", type="primary", use_container_width=True):
            try:
                sb: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
                res = sb.table("usuarios_app").select("*").eq("email", email.lower()).execute()
                if not res.data:
                    st.error("Usuario no encontrado.")
                else:
                    user_db = res.data[0]
                    if user_db["password"] == hash_password(passwd):
                        if user_db["is_active"]:
                            st.session_state.logged_in_user = user_db
                            st.session_state.show_login_modal = False
                            st.rerun()
                        else:
                            st.warning("Cuenta registrada pero aún no habilitada por el administrador.")
                    else:
                        st.error("Contraseña incorrecta.")
            except Exception as e:
                st.error(f"Error de conexión: {e}")

    with tab_registro:
        nombre = st.text_input("Nombre", key="reg_nombre")
        apellido = st.text_input("Apellido", key="reg_apellido")
        empresa = st.text_input("Empresa", key="reg_empresa")
        email_r = st.text_input("Correo Electrónico", key="reg_email")
        p1 = st.text_input("Contraseña", type="password", key="reg_p1")
        p2 = st.text_input("Confirmar Contraseña", type="password", key="reg_p2")
        if st.button("Registrarse", type="primary", use_container_width=True, key="btn_reg"):
            if not all([nombre, apellido, empresa, email_r, p1, p2]):
                st.warning("Complete todos los campos.")
            elif p1 != p2:
                st.error("Las contraseñas no coinciden.")
            elif len(p1) < 6:
                st.error("La contraseña debe tener al menos 6 caracteres.")
            else:
                try:
                    sb: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
                    check = sb.table("usuarios_app").select("email").eq("email", email_r.lower()).execute()
                    if check.data:
                        st.error("Este correo ya está registrado.")
                    else:
                        sb.table("usuarios_app").insert({
                            "nombre": nombre, "apellido": apellido, "empresa": empresa,
                            "email": email_r.lower(), "password": hash_password(p1), "is_active": False
                        }).execute()
                        st.success("✅ Registro completado. Espera la habilitación del administrador.")
                except Exception as e:
                    st.error(f"Error al registrar: {e}")

# ==============================================================================
# MÓDULO DE CREDENCIALES API (Admin only)
# ==============================================================================
@st.dialog("⚙️ Configuración Segura — APIs", width="small")
def modal_credenciales():
    st.caption("Ingresa el PIN de administrador para editar.")
    pin = st.text_input("PIN Secreto", type="password", key="cred_pin")

    if st.button("Verificar PIN", use_container_width=True):
        try:
            sb: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
            r = sb.table("app_settings").select("setting_value").eq("setting_name", "admin_code").execute()
            if r.data and r.data[0]["setting_value"] == pin:
                st.session_state.admin_unlocked = True
                st.success("✅ Identidad verificada. Ya puedes editar.")
            else:
                st.error("PIN incorrecto.")
        except Exception as e:
            st.error(f"Error de conexión: {e}")

    if st.session_state.admin_unlocked:
        GEMINI_KEY, GEMINI_BACKUP, DEEPSEEK_KEY, _ = get_all_api_keys()

        st.divider()
        col1, col2 = st.columns([3, 1])
        with col1:
            st.markdown("**Gemini Principal**")
        with col2:
            st.session_state.api_1_enabled = st.checkbox("Activo", value=st.session_state.api_1_enabled, key="chk_api1")
        nueva_api1 = st.text_input("API Key", value=GEMINI_KEY or "", key="edit_api1")

        col3, col4 = st.columns([3, 1])
        with col3:
            st.markdown("**Gemini Backup**")
        with col4:
            st.session_state.api_2_enabled = st.checkbox("Activo", value=st.session_state.api_2_enabled, key="chk_api2")
        nueva_api2 = st.text_input("API Key", value=GEMINI_BACKUP or "", key="edit_api2")

        col5, col6 = st.columns([3, 1])
        with col5:
            st.markdown("**DeepSeek**")
        with col6:
            st.session_state.api_3_enabled = st.checkbox("Activo", value=st.session_state.api_3_enabled, key="chk_api3")
        nueva_api3 = st.text_input("API Key", value=DEEPSEEK_KEY or "", key="edit_api3")

        if st.button("💾 Guardar Claves", type="primary", use_container_width=True):
            try:
                sb: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
                if nueva_api1:
                    sb.table("app_settings").upsert({"setting_name": "gemini_api_key", "setting_value": nueva_api1}).execute()
                if nueva_api2:
                    sb.table("app_settings").upsert({"setting_name": "gemini_api_key_backup", "setting_value": nueva_api2}).execute()
                if nueva_api3:
                    sb.table("app_settings").upsert({"setting_name": "deepseek_api_key", "setting_value": nueva_api3}).execute()
                st.success("✅ Claves guardadas en Supabase.")
            except Exception as e:
                st.error(f"Error al guardar: {e}")

# ==============================================================================
# UI PRINCIPAL
# ==============================================================================
def main():
    # --- SIDEBAR (Siempre renderizada con algo para evitar que desaparezca) ---
    with st.sidebar:
        # Logo DESPUX con clase personalizada
        st.markdown('<a href="https://www.despux.net/" target="_blank" class="despux-logo">DESPUX</a>', unsafe_allow_html=True)
        st.markdown("---")
        
        # Heading con clase para visibilidad
        st.markdown('<div class="sb-heading">📂 Documentos</div>', unsafe_allow_html=True)

        # Botón PDF (Azul original)
        st.markdown("""
            <style>
            div[data-testid="stFileUploader"] > section > button { background-color: #007AFF !important; color: white !important; }
            </style>
        """, unsafe_allow_html=True)
        pdf_file = st.file_uploader("📄 Subir Oferta (PDF)", type=["pdf"], key="pdf_uploader", help="Selecciona el PDF principal del proyecto")
        
        # Botón Extras (Verde original)
        st.markdown("""
            <style>
            /* Hack para colorear el segundo uploader si es posible, o simplemente forzar visibilidad */
            [data-testid="stFormSubmitButton"] > button { background-color: #34C759 !important; }
            </style>
        """, unsafe_allow_html=True)
        extra_files = st.file_uploader("📝 Archivos Extras (.txt, .docx)", type=["txt", "docx"], accept_multiple_files=True, key="extra_uploader")

        # Botón GENERAR BITÁCORA (Naranja original)
        generar_disabled = not st.session_state.processing_done
        st.markdown(f"""
            <style>
            button[key="btn_generar"] {{ background-color: #FF9500 !important; color: white !important; height: 50px !important; font-size: 16px !important; }}
            </style>
        """, unsafe_allow_html=True)
        generar_clicked = st.button("📋 GENERAR BITÁCORA", use_container_width=True, disabled=generar_disabled, key="btn_generar")

        st.markdown("---")
        sidebar_auth()
        st.markdown("---")
        
        # Botón Configuración (Gris suave original)
        if st.button("⚙️ Configuración Segura", use_container_width=True, key="btn_config"):
            st.session_state.show_cred_modal = True
            st.rerun()

    # --- MODALES ---
    if st.session_state.get("show_login_modal"):
        st.session_state.show_login_modal = False
        modal_login()

    if st.session_state.get("show_cred_modal"):
        st.session_state.show_cred_modal = False
        modal_credenciales()

    # --- CONTENIDO PRINCIPAL ---
    st.markdown("## 🚀 Bot de Bitácora")
    st.caption("Procesamiento inteligente de Google Sheets con IA · DESPUX")
    st.divider()

    col_inputs, col_resultados = st.columns([1, 1.8], gap="medium")

    with col_inputs:
        with st.container(border=True):
            st.markdown("### 📝 Datos Manuales")
            factura = st.text_input("Número de Factura", placeholder="Ej: FAC-2026-0012")
            orden = st.text_input("Número de Orden de Compra", placeholder="Ej: OC-459821")
            fecha_ini = st.text_input("Fecha de Inicio del Proyecto", placeholder="Ej: 22 de Marzo de 2026")
            fecha_fin = st.text_input("Fecha de Finalización", placeholder="Ej: 30 de Noviembre de 2026")
            herramientas = st.text_input("Herramientas de Medición", placeholder="Ej: Multímetro, Vernier")
            descripcion = st.text_area("Descripción Detallada", height=80)
            hitos = st.text_area("Hitos Principales", height=80)

            st.markdown("---")
            procesar_disabled = st.session_state.logged_in_user is None
            if procesar_disabled:
                st.info("🔒 Inicia sesión para procesar.")

            col_proc, col_stop = st.columns(2)
            with col_proc:
                # Botón Procesar (Azul)
                procesar = st.button("🚀 Procesar", use_container_width=True, disabled=procesar_disabled, key="btn_procesar")
            with col_stop:
                # Botón Detener (Rojo)
                st.button("🛑 Detener", use_container_width=True, disabled=True, key="btn_detener")

    with col_resultados:
        with st.container(border=True):
            st.markdown("### 📊 Resultados Generados")

            # Barra de progreso
            progress_bar = st.progress(0)
            lbl_progreso = st.empty()

            # Tabla de resultados en vivo
            tabla_placeholder = st.empty()
            if st.session_state.datos_procesados:
                tabla_placeholder.dataframe(
                    pd.DataFrame(st.session_state.datos_procesados),
                    use_container_width=True,
                    height=250
                )
            else:
                tabla_placeholder.info("Los resultados aparecerán aquí durante el procesamiento.")

            # Botones de acción post-proceso
            if st.session_state.datos_procesados:
                col_exp, col_lmp = st.columns(2)
                with col_exp:
                    # Exportar a Excel (Naranja)
                    df_exp = pd.DataFrame(st.session_state.datos_procesados)
                    buf = io.BytesIO()
                    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                        df_exp.to_excel(writer, index=False, sheet_name="Resultados")
                    st.download_button(
                        "📥 Exportar a Excel",
                        data=buf.getvalue(),
                        file_name="bitacora_resultados.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                with col_lmp:
                    # Limpiar Datos (Rojo)
                    if st.button("🗑 Limpiar Datos", use_container_width=True, key="btn_limpiar"):
                        st.session_state.datos_procesados = []
                        st.session_state.log_lines = []
                        st.session_state.processing_done = False
                        st.rerun()

        st.markdown(" ") # Espacio
        
        with st.container(border=True):
            st.markdown("### 💻 Terminal de Sistema")
            log_placeholder = st.empty()
            if st.session_state.log_lines:
                render_log(log_placeholder)
            else:
                st.caption("Esperando inicio de proceso...")

    # --- GENERAR BITÁCORA ---
    if generar_clicked and pdf_file:
        st.info("Funcionalidad 'GENERAR BITÁCORA': selecciona un archivo Excel al que deseas agregar una hoja 'despux' con los resultados.")
        excel_upload = st.file_uploader("Sube el archivo Excel destino", type=["xlsx"], key="excel_dest")
        if excel_upload and st.session_state.datos_procesados:
            try:
                wb = openpyxl.load_workbook(io.BytesIO(excel_upload.read()))
                if "despux" in wb.sheetnames:
                    ws = wb["despux"]
                    ws.delete_rows(1, ws.max_row)
                else:
                    ws = wb.create_sheet("despux")
                ws.append(["Celda", "Etiqueta (Col A)", "Valor Generado (Col B)", "Hoja"])
                for row in st.session_state.datos_procesados:
                    ws.append([row["Celda"], row["Col A (Etiqueta)"], row["Col B (Generado)"], row["Hoja"]])
                buf_wb = io.BytesIO()
                wb.save(buf_wb)
                st.download_button("📥 Descargar Excel con hoja 'despux'", data=buf_wb.getvalue(),
                                   file_name="bitacora_despux.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                st.success("✅ Hoja 'despux' generada correctamente.")
            except Exception as e:
                st.error(f"Error al generar la bitácora: {e}")

    # --- LANZAR PROCESAMIENTO ---
    if procesar and pdf_file:
        datos_usuario = {
            "Factura": factura, "Orden de Compra": orden,
            "Fecha Inicio": fecha_ini, "Fecha Finalización": fecha_fin,
            "Herramientas de Medición": herramientas,
            "Descripción del Proyecto": descripcion,
            "Hitos Principales": hitos,
        }
        procesar_todo(
            pdf_bytes=pdf_file.read(),
            archivos_extras=extra_files,
            datos_usuario=datos_usuario,
            progress_bar=progress_bar,
            lbl_progreso=lbl_progreso,
            tabla_placeholder=tabla_placeholder,
            log_placeholder=log_placeholder,
        )
        st.rerun()
    elif procesar and not pdf_file:
        st.warning("⚠️ Debes subir un PDF antes de procesar.")

if __name__ == "__main__":
    main()
